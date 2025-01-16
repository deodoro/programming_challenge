import openpyxl
import csv

df = openpyxl.load_workbook("nsdp_delays_random.xlsx", data_only=True)
worksheet = df.get_sheet_by_name("Sheet1")
output_file_name = "output.csv"

r_index, c_index = 2,2 
t = c_index
total = 0 
while True:
    while True:
        value = worksheet.cell(row=r_index, column=t).value
        t+=1
        if value is None:
            break
    total = t - 1
    break

r_index+=1
data_points = {}
years = set()
while True:
    t = 8

    timely = worksheet.cell(row=r_index, column=4).value
    if timely is None:
        break
    if timely.lower() != 't':
        r_index+=1
        continue
    year = worksheet.cell(row=r_index, column=3).value
    country = worksheet.cell(row=r_index, column=2).value
    years.add(year)
    if country not in data_points:
        data_points[country] ={year:  {'total':0, 'counts':0}}
    if year not in data_points[country]:
        data_points[country][year] = {'total':0, 'counts':0}

    while t <= total:
        value = worksheet.cell(row=r_index, column=t).value
        t+=1
        try:
            value = int(value)
            if value > 0:
                data_points[country][year]['total']+=value
        except:
            pass 
    data_points[country][year]['counts']+=1       
    total = t - 1
    r_index+=1

with open(output_file_name ,'w', newline='') as csvfile:
    csvWriter =  csv.writer(csvfile, delimiter=',')
    years = sorted(list(years))

    fieldnames = ["country"] + years
    # create a CSV writer object
    csvWriter.writerow(fieldnames)

    rows = []
    for country, y_v in data_points.items():
        row = [country]
        for year in years:
            if year in y_v:
                row.append( y_v[year]['total'] /  y_v[year]['counts']  )
            else:
                row.append('NA')
        rows.append(row)

    for row in rows:
        csvWriter.writerow(row)

    

